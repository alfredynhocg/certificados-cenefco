import hashlib
import io
import re
import shutil
import unicodedata
import zipfile
from pathlib import Path

import fitz
import openpyxl
import qrcode
from PIL import Image
from reportlab.lib.colors import HexColor
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfgen import canvas

class GeneradorError(Exception):
    pass


def limpiar_nombre(valor) -> str | None:
    if valor is None:
        return None
    texto = str(valor).strip()
    texto = re.sub(r"\s+", " ", texto)
    texto = texto.strip(" ,.-:;")
    return texto or None


def nombre_archivo_seguro(nombre: str) -> str:
    sin_acentos = unicodedata.normalize("NFKD", nombre).encode("ascii", "ignore").decode("ascii")
    limpio = re.sub(r"[^A-Za-z0-9 _-]", "", sin_acentos).strip()
    limpio = re.sub(r"\s+", "_", limpio)
    return limpio or "certificado"


def normalizar_busqueda(texto) -> str:
    if not texto:
        return ""
    sin_acentos = unicodedata.normalize("NFKD", str(texto)).encode("ascii", "ignore").decode("ascii")
    return sin_acentos.lower().strip()


def leer_nombres(ruta_excel) -> list[str]:
    wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    ws = wb.active

    nombres = []
    vistos = set()
    for fila in ws.iter_rows(min_row=1, max_col=1):
        nombre = limpiar_nombre(fila[0].value)
        if nombre is None:
            continue
        clave = nombre.lower()
        if clave in vistos:
            continue
        vistos.add(clave)
        nombres.append(nombre)
    return nombres


def filas_desde_excel(ruta_excel) -> list[dict]:
    wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    ws = wb.active

    filas = []
    orden = 0
    for fila in ws.iter_rows(min_row=1, max_col=2):
        valor_original = fila[0].value
        nombre = limpiar_nombre(valor_original)
        fecha_registro = fila[1].value
        if nombre is None and (valor_original is None or str(valor_original).strip() == ""):
            continue
        orden += 1
        filas.append({
            "orden": orden,
            "valor_original": "" if valor_original is None else str(valor_original),
            "nombre": nombre or "",
            "fecha_registro": fecha_registro if hasattr(fecha_registro, "year") else None,
        })
    return filas


def poblar_estudiantes_desde_excel(lote, ruta_excel=None) -> int:
    from .models import Estudiante

    ruta = ruta_excel if ruta_excel is not None else Path(lote.excel.path)
    filas = filas_desde_excel(ruta)

    lote.estudiantes.all().delete()
    Estudiante.objects.bulk_create([
        Estudiante(
            lote=lote,
            orden=fila["orden"],
            valor_original=fila["valor_original"],
            nombre=fila["nombre"],
            nombre_normalizado=normalizar_busqueda(fila["nombre"] or fila["valor_original"]),
            fecha_registro=fila["fecha_registro"],
        )
        for fila in filas
    ])
    return len(filas)


PATRON_CARACTERES_RAROS = re.compile(r"[^A-Za-zÁÉÍÓÚÑÜáéíóúñü\s'.-]")
PATRON_SOLO_LETRAS = re.compile(r"[A-Za-zÁÉÍÓÚÑÜáéíóúñü]")


def analizar_estudiantes(lote, fecha_por_defecto=None) -> dict:
    filas = []
    vistos = {}
    nombres_finales = []
    nombres_vistos = set()

    for estudiante in lote.estudiantes.all():
        valor_original = estudiante.valor_original
        nombre = limpiar_nombre(estudiante.nombre) or limpiar_nombre(valor_original)
        fecha_registro = estudiante.fecha_registro or fecha_por_defecto

        advertencias = []
        if nombre is None:
            if valor_original.strip():
                advertencias.append("Fila vacia tras limpiar el texto")
            else:
                continue
        else:
            if len(nombre) < 4:
                advertencias.append("Nombre muy corto, revisar si esta incompleto")
            if not PATRON_SOLO_LETRAS.search(nombre):
                advertencias.append("No contiene letras")
            caracteres_raros = sorted(set(PATRON_CARACTERES_RAROS.findall(nombre)))
            if caracteres_raros:
                advertencias.append(f"Contiene caracteres inusuales: {' '.join(caracteres_raros)}")
            clave = nombre.lower()
            if clave in vistos:
                advertencias.append(f"Duplicado de la fila {vistos[clave]}")
            else:
                vistos[clave] = estudiante.orden
                if clave not in nombres_vistos:
                    nombres_vistos.add(clave)
                    nombres_finales.append(nombre)

        filas.append({
            "fila": estudiante.orden,
            "valor_original": valor_original,
            "nombre": nombre,
            "advertencias": advertencias,
            "fecha_registro": fecha_registro,
        })

    return {
        "filas": filas,
        "total_filas_con_datos": len(filas),
        "total_nombres_validos": len(nombres_finales),
        "filas_con_advertencias": [f for f in filas if f["advertencias"]],
    }


def nombres_desde_estudiantes(lote) -> list[str]:
    nombres = []
    vistos = set()
    for estudiante in lote.estudiantes.all():
        nombre = limpiar_nombre(estudiante.nombre) or limpiar_nombre(estudiante.valor_original)
        if nombre is None:
            continue
        clave = nombre.lower()
        if clave in vistos:
            continue
        vistos.add(clave)
        nombres.append(nombre)
    return nombres


def exportar_reporte_excel(analisis: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"

    encabezados = ["Fila", "Valor original", "Nombre que se usara", "Fecha de registro", "Advertencias"]
    ws.append(encabezados)
    for celda in ws[1]:
        celda.font = openpyxl.styles.Font(bold=True)

    relleno_advertencia = openpyxl.styles.PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

    for item in analisis["filas"]:
        fecha = item.get("fecha_registro")
        fila_valores = [
            item["fila"],
            item["valor_original"],
            item["nombre"] or "(vacio)",
            fecha.strftime("%d/%m/%Y") if hasattr(fecha, "strftime") else (fecha or ""),
            "; ".join(item["advertencias"]),
        ]
        ws.append(fila_valores)
        if item["advertencias"]:
            for celda in ws[ws.max_row]:
                celda.fill = relleno_advertencia

    for columna, ancho in zip("ABCDE", (8, 35, 35, 18, 50)):
        ws.column_dimensions[columna].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def calcular_tamano_fuente(texto: str, fuente: str, tamano_inicial: int, tamano_min: int, ancho_max_pt: float) -> int:
    tamano = tamano_inicial
    while tamano > tamano_min and stringWidth(texto, fuente, tamano) > ancho_max_pt:
        tamano -= 1
    return tamano


def preparar_plantilla_comprimida(template_path: Path, carpeta_temporal: Path, calidad: int = 85) -> Path:
    carpeta_temporal.mkdir(parents=True, exist_ok=True)
    destino = carpeta_temporal / "plantilla_comprimida.jpg"
    with Image.open(template_path) as img:
        img.convert("RGB").save(destino, "JPEG", quality=calidad, optimize=True)
    return destino


def generar_imagen_qr(contenido: str) -> ImageReader:
    qr = qrcode.QRCode(border=1)
    qr.add_data(contenido)
    qr.make(fit=True)
    imagen = qr.make_image(fill_color="black", back_color="white").convert("RGB")
    return ImageReader(imagen)


def _dibujar_certificado(c: canvas.Canvas, nombre: str, template_path: Path, ancho_pt: int, alto_pt: int, opciones: dict):
    c.drawImage(str(template_path), 0, 0, width=ancho_pt, height=alto_pt)

    fuente = opciones.get("fuente", "Helvetica-Bold")
    color = opciones.get("color", "#1a1a1a")
    tamano_max = opciones.get("tamano_fuente", 40)
    tamano_min = opciones.get("tamano_fuente_min", 18)

    x_px = opciones["x"]
    y_px = opciones["y"]
    y_pdf = alto_pt - y_px

    margen_izq = opciones.get("margen_izquierdo", 0)
    margen_der = opciones.get("margen_derecho", 0)
    espacio_izq = x_px - margen_izq
    espacio_der = (ancho_pt - margen_der) - x_px
    ancho_max = 2 * min(espacio_izq, espacio_der)

    tamano = calcular_tamano_fuente(nombre, fuente, tamano_max, tamano_min, max(ancho_max, 1))

    c.setFont(fuente, tamano)
    c.setFillColor(HexColor(color))
    c.drawCentredString(x_px, y_pdf, nombre)

    qr_imagen = opciones.get("qr_imagen")
    if qr_imagen is not None:
        qr_tamano = opciones.get("qr_tamano", 120)
        qr_x_pdf = opciones["qr_x"]
        qr_y_pdf = alto_pt - opciones["qr_y"]
        c.drawImage(
            qr_imagen,
            qr_x_pdf - qr_tamano / 2,
            qr_y_pdf - qr_tamano / 2,
            width=qr_tamano,
            height=qr_tamano,
        )


def generar_codigo_certificado(lote_pk: int, nombre: str) -> str:
    """Codigo corto y estable, unico por participante dentro de un lote."""
    base = f"{lote_pk}:{nombre.strip().lower()}"
    return hashlib.sha256(base.encode("utf-8")).hexdigest()[:10].upper()


def renderizar_contenido_qr(plantilla: str, *, nombre: str, curso: str, codigo: str) -> str:
    return (
        plantilla
        .replace("{nombre}", nombre)
        .replace("{curso}", curso)
        .replace("{codigo}", codigo)
    )


def _resolver_opciones_dibujo(opciones: dict, nombre: str) -> dict:
    resueltas = dict(opciones)
    qr_plantilla = resueltas.pop("qr_plantilla", None)
    curso = resueltas.pop("curso", "")
    lote_pk = resueltas.pop("lote_pk", 0)
    if qr_plantilla:
        codigo = generar_codigo_certificado(lote_pk, nombre)
        contenido = renderizar_contenido_qr(qr_plantilla, nombre=nombre, curso=curso, codigo=codigo)
        resueltas["qr_imagen"] = generar_imagen_qr(contenido)
    return resueltas


def generar_pdf_certificado(nombre: str, pdf_path: Path, template_path: Path, img_w_px: int, img_h_px: int, opciones: dict):
    opciones_dibujo = _resolver_opciones_dibujo(opciones, nombre)
    c = canvas.Canvas(str(pdf_path), pagesize=(img_w_px, img_h_px))
    _dibujar_certificado(c, nombre, template_path, img_w_px, img_h_px, opciones_dibujo)
    c.showPage()
    c.save()


def opciones_desde_lote(lote) -> dict:
    opciones = {
        "x": lote.texto_x,
        "y": lote.texto_y,
        "margen_izquierdo": lote.margen_izquierdo,
        "margen_derecho": lote.margen_derecho,
        "fuente": lote.fuente,
        "tamano_fuente": lote.tamano_fuente,
        "tamano_fuente_min": lote.tamano_fuente_min,
        "color": lote.color_texto,
    }
    if lote.qr_activo and lote.qr_contenido and lote.qr_x is not None and lote.qr_y is not None:
        opciones["qr_plantilla"] = lote.qr_contenido
        opciones["qr_x"] = lote.qr_x
        opciones["qr_y"] = lote.qr_y
        opciones["qr_tamano"] = lote.qr_tamano
        opciones["curso"] = lote.curso
        opciones["lote_pk"] = lote.pk
    return opciones


def generar_previsualizacion_png(lote, ancho_max_px: int = 1000) -> bytes:
    template_path = Path(lote.plantilla.path)
    if not template_path.exists():
        raise GeneradorError("No se encontro el archivo de plantilla.")

    with Image.open(template_path) as img:
        img_w_px, img_h_px = img.size

    nombre_muestra = "Nombre de Ejemplo Apellido"
    try:
        nombres = nombres_desde_estudiantes(lote)
        if nombres:
            nombre_muestra = nombres[0]
    except Exception:
        pass

    opciones_dibujo = _resolver_opciones_dibujo(opciones_desde_lote(lote), nombre_muestra)

    buffer_pdf = io.BytesIO()
    c = canvas.Canvas(buffer_pdf, pagesize=(img_w_px, img_h_px))
    _dibujar_certificado(c, nombre_muestra, template_path, img_w_px, img_h_px, opciones_dibujo)
    c.showPage()
    c.save()
    buffer_pdf.seek(0)

    doc = fitz.open(stream=buffer_pdf.read(), filetype="pdf")
    zoom = min(ancho_max_px / img_w_px, 2.0)
    pix = doc[0].get_pixmap(matrix=fitz.Matrix(zoom, zoom))
    return pix.tobytes("png")


def contar_nombres(lote) -> int:
    return len(nombres_desde_estudiantes(lote))


def generar_certificados_lote(lote, carpeta_salida: Path, on_progreso=None) -> tuple[int, Path]:
    template_path = Path(lote.plantilla.path)

    if not template_path.exists():
        raise GeneradorError("No se encontro el archivo de plantilla.")

    with Image.open(template_path) as img:
        img_w_px, img_h_px = img.size

    nombres = nombres_desde_estudiantes(lote)
    if not nombres:
        raise GeneradorError("El lote no tiene participantes registrados.")

    carpeta_salida.mkdir(parents=True, exist_ok=True)

    carpeta_temporal = carpeta_salida.parent / "_tmp"
    plantilla_para_pdf = preparar_plantilla_comprimida(template_path, carpeta_temporal)

    opciones = opciones_desde_lote(lote)

    total = len(nombres)
    pdfs = []
    usados = {}
    try:
        for indice, nombre in enumerate(nombres, start=1):
            base = nombre_archivo_seguro(nombre)
            usados[base] = usados.get(base, 0) + 1
            sufijo = "" if usados[base] == 1 else f"_{usados[base]}"
            pdf_path = carpeta_salida / f"{base}{sufijo}.pdf"
            generar_pdf_certificado(nombre, pdf_path, plantilla_para_pdf, img_w_px, img_h_px, opciones)
            pdfs.append(pdf_path)
            if on_progreso:
                on_progreso(indice, total)

        zip_path = carpeta_salida / "certificados.zip"
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for pdf_path in pdfs:
                zf.write(pdf_path, arcname=pdf_path.name)
    finally:
        shutil.rmtree(carpeta_temporal, ignore_errors=True)

    return len(pdfs), zip_path


def nombre_archivo_para(nombres: list[str], nombre_objetivo: str) -> str | None:
    """Replica la desambiguacion de nombres duplicados de generar_certificados_lote,
    para saber que nombre de archivo le tocaria a un nombre puntual dentro de la lista completa."""
    usados = {}
    resultado = None
    for nombre in nombres:
        base = nombre_archivo_seguro(nombre)
        usados[base] = usados.get(base, 0) + 1
        sufijo = "" if usados[base] == 1 else f"_{usados[base]}"
        if nombre == nombre_objetivo and resultado is None:
            resultado = f"{base}{sufijo}"
    return resultado


def generar_certificado_individual(lote, nombre: str, carpeta_salida: Path) -> Path:
    """Genera (o regenera) el PDF de un unico participante, sin tocar los demas
    certificados ya generados en el lote."""
    template_path = Path(lote.plantilla.path)
    if not template_path.exists():
        raise GeneradorError("No se encontro el archivo de plantilla.")

    nombres = nombres_desde_estudiantes(lote)
    if nombre not in nombres:
        raise GeneradorError("El participante ya no esta en la lista del lote.")

    nombre_archivo = nombre_archivo_para(nombres, nombre)

    with Image.open(template_path) as img:
        img_w_px, img_h_px = img.size

    carpeta_salida.mkdir(parents=True, exist_ok=True)
    carpeta_temporal = carpeta_salida.parent / "_tmp"
    plantilla_para_pdf = preparar_plantilla_comprimida(template_path, carpeta_temporal)

    try:
        opciones = opciones_desde_lote(lote)
        pdf_path = carpeta_salida / f"{nombre_archivo}.pdf"
        generar_pdf_certificado(nombre, pdf_path, plantilla_para_pdf, img_w_px, img_h_px, opciones)
    finally:
        shutil.rmtree(carpeta_temporal, ignore_errors=True)

    return pdf_path
