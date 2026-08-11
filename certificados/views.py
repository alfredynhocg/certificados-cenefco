import io
import shutil
import threading
import zipfile
from pathlib import Path

import openpyxl
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.files import File
from django.db import close_old_connections
from django.db.models import Count, Max, Q, Sum
from django.db.models.functions import TruncMonth
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone

from .forms import LoteForm
from .generador import (
    GeneradorError,
    analizar_estudiantes,
    exportar_reporte_excel,
    generar_certificado_individual as generar_certificado_individual_pdf,
    generar_certificados_lote,
    generar_previsualizacion_png,
    limpiar_nombre,
    nombre_archivo_seguro,
    normalizar_busqueda,
    nombres_desde_estudiantes,
    poblar_estudiantes_desde_excel,
)
from .models import Estudiante, Lote


def _puede_ver(user, lote) -> bool:
    return user.es_admin or lote.propietario_id == user.id


def _mapa_certificados(lote) -> dict:
    carpeta_salida = Path(settings.MEDIA_ROOT) / "lotes" / str(lote.pk) / "salida"
    if not carpeta_salida.exists():
        return {}
    try:
        nombres = nombres_desde_estudiantes(lote)
    except Exception:
        return {}

    usados = {}
    mapa = {}
    for nombre in nombres:
        base = nombre_archivo_seguro(nombre)
        usados[base] = usados.get(base, 0) + 1
        sufijo = "" if usados[base] == 1 else f"_{usados[base]}"
        ruta = carpeta_salida / f"{base}{sufijo}.pdf"
        if ruta.exists():
            mapa[nombre] = ruta
    return mapa


@login_required
def dashboard(request):
    if request.user.es_admin:
        lotes = Lote.objects.select_related("propietario").all()
    else:
        lotes = Lote.objects.filter(propietario=request.user)

    estado = request.GET.get("estado", "").strip()
    if estado in Lote.Estado.values:
        lotes = lotes.filter(estado=estado)

    desde = request.GET.get("desde", "").strip()
    if desde:
        lotes = lotes.filter(creado_en__date__gte=desde)

    hasta = request.GET.get("hasta", "").strip()
    if hasta:
        lotes = lotes.filter(creado_en__date__lte=hasta)

    lotes = list(lotes)

    busqueda = request.GET.get("q", "").strip()
    coincidencias_participante = {}
    if busqueda:
        termino = normalizar_busqueda(busqueda)
        ids_lotes_visibles = [lote.id for lote in lotes]

        nombres_por_lote = {}
        estudiantes_candidatos = Estudiante.objects.filter(
            lote_id__in=ids_lotes_visibles, nombre_normalizado__icontains=termino
        ).order_by("lote_id", "orden")
        for estudiante in estudiantes_candidatos:
            nombres_por_lote.setdefault(estudiante.lote_id, []).append(estudiante.nombre)

        ids_coincidentes = []
        for lote in lotes:
            coincide = termino in normalizar_busqueda(lote.curso)
            if not coincide and request.user.es_admin:
                coincide = termino in normalizar_busqueda(lote.propietario.username)

            nombres_coincidentes = nombres_por_lote.get(lote.id, [])
            if nombres_coincidentes:
                mapa_certificados = _mapa_certificados(lote)
                coincidencias_participante[lote.id] = [
                    {"nombre": n, "generado": n in mapa_certificados} for n in nombres_coincidentes
                ]

            if coincide or nombres_coincidentes:
                ids_coincidentes.append(lote.id)

        lotes = [lote for lote in lotes if lote.id in set(ids_coincidentes)]

    cursos_con_participante = []
    cursos_sin_certificado = []
    for lote in lotes:
        info = coincidencias_participante.get(lote.id, [])
        lote.coincidencias_nombre = info
        if not info:
            continue
        if any(item["generado"] for item in info):
            cursos_con_participante.append(lote.curso)
        else:
            cursos_sin_certificado.append(lote.curso)

    return render(request, "certificados/dashboard.html", {
        "lotes": lotes,
        "busqueda": busqueda,
        "estado_seleccionado": estado,
        "estados": Lote.Estado.choices,
        "desde": desde,
        "hasta": hasta,
        "coincidencias_participante": coincidencias_participante,
        "cursos_con_participante": cursos_con_participante,
        "cursos_sin_certificado": cursos_sin_certificado,
    })


ESTADOS_KPI = [
    (Lote.Estado.GENERADO, "bg-emerald-500", "#10b981"),
    (Lote.Estado.PROCESANDO, "bg-amber-500", "#f59e0b"),
    (Lote.Estado.ERROR, "bg-red-500", "#ef4444"),
    (Lote.Estado.BORRADOR, "bg-slate-400", "#94a3b8"),
]


def _calcular_estadisticas(user) -> dict:
    if user.es_admin:
        lotes = Lote.objects.all()
    else:
        lotes = Lote.objects.filter(propietario=user)

    total_lotes = lotes.count()
    generados = lotes.filter(estado=Lote.Estado.GENERADO)
    total_lotes_generados = generados.count()
    total_certificados = generados.aggregate(total=Sum("total_certificados"))["total"] or 0
    promedio_por_lote = round(total_certificados / total_lotes_generados, 1) if total_lotes_generados else 0

    conteo_por_estado = {fila["estado"]: fila["cantidad"] for fila in lotes.values("estado").annotate(cantidad=Count("id"))}
    estados_resumen = []
    for valor, color, color_hex in ESTADOS_KPI:
        cantidad = conteo_por_estado.get(valor, 0)
        estados_resumen.append({
            "etiqueta": Lote.Estado(valor).label,
            "cantidad": cantidad,
            "color": color,
            "color_hex": color_hex,
            "porcentaje": round((cantidad / total_lotes) * 100) if total_lotes else 0,
        })

    por_curso_qs = list(
        generados.values("curso").annotate(cantidad=Sum("total_certificados")).order_by("-cantidad")
    )
    total_tipos_curso = len(por_curso_qs)
    por_curso = por_curso_qs[:10]
    max_por_curso = por_curso[0]["cantidad"] if por_curso else 0
    for item in por_curso:
        item["porcentaje"] = round((item["cantidad"] / max_por_curso) * 100) if max_por_curso else 0

    tendencia = list(
        generados.exclude(generado_en=None)
        .annotate(mes=TruncMonth("generado_en"))
        .values("mes")
        .annotate(cantidad=Sum("total_certificados"))
        .order_by("mes")
    )[-6:]
    max_tendencia = max((item["cantidad"] for item in tendencia), default=0)
    for item in tendencia:
        item["porcentaje"] = round((item["cantidad"] / max_tendencia) * 100) if max_tendencia else 0

    top_usuarios = []
    if user.es_admin:
        top_usuarios = list(
            generados.values("propietario__username")
            .annotate(cantidad=Sum("total_certificados"))
            .order_by("-cantidad")[:8]
        )
        max_usuario = top_usuarios[0]["cantidad"] if top_usuarios else 0
        for item in top_usuarios:
            item["porcentaje"] = round((item["cantidad"] / max_usuario) * 100) if max_usuario else 0

    total_participantes = Estudiante.objects.filter(lote__in=lotes).count()

    por_curso_participantes_qs = list(
        Estudiante.objects.filter(lote__in=lotes)
        .values("lote__curso")
        .annotate(cantidad=Count("id"))
        .order_by("-cantidad")[:10]
    )
    max_participantes_curso = por_curso_participantes_qs[0]["cantidad"] if por_curso_participantes_qs else 0
    for item in por_curso_participantes_qs:
        item["porcentaje"] = round((item["cantidad"] / max_participantes_curso) * 100) if max_participantes_curso else 0

    return {
        "total_lotes": total_lotes,
        "total_lotes_generados": total_lotes_generados,
        "total_certificados": total_certificados,
        "promedio_por_lote": promedio_por_lote,
        "estados_resumen": estados_resumen,
        "por_curso": por_curso,
        "total_tipos_curso": total_tipos_curso,
        "tendencia": tendencia,
        "top_usuarios": top_usuarios,
        "altura_grafico_curso": max(len(por_curso) * 32, 190),
        "altura_grafico_usuarios": max(len(top_usuarios) * 32, 190),
        "total_participantes": total_participantes,
        "por_curso_participantes": por_curso_participantes_qs,
        "altura_grafico_participantes": max(len(por_curso_participantes_qs) * 32, 190),
    }


@login_required
def estadisticas(request):
    return render(request, "certificados/estadisticas.html", _calcular_estadisticas(request.user))


@login_required
def exportar_estadisticas_excel(request):
    datos = _calcular_estadisticas(request.user)

    wb = openpyxl.Workbook()
    negrita = openpyxl.styles.Font(bold=True)

    ws = wb.active
    ws.title = "Resumen"
    ws.append(["Indicador", "Valor"])
    for celda in ws[1]:
        celda.font = negrita
    ws.append(["Certificados emitidos", datos["total_certificados"]])
    ws.append(["Participantes inscritos", datos["total_participantes"]])
    ws.append(["Tipos de certificado (cursos distintos)", datos["total_tipos_curso"]])
    ws.append(["Lotes generados", datos["total_lotes_generados"]])
    ws.append(["Total de lotes", datos["total_lotes"]])
    ws.append(["Promedio de certificados por lote", datos["promedio_por_lote"]])
    ws.column_dimensions["A"].width = 38

    ws_estados = wb.create_sheet("Lotes por estado")
    ws_estados.append(["Estado", "Cantidad", "Porcentaje"])
    for celda in ws_estados[1]:
        celda.font = negrita
    for item in datos["estados_resumen"]:
        ws_estados.append([item["etiqueta"], item["cantidad"], f'{item["porcentaje"]}%'])
    ws_estados.column_dimensions["A"].width = 20

    ws_curso = wb.create_sheet("Certificados por curso")
    ws_curso.append(["Curso", "Certificados emitidos"])
    for celda in ws_curso[1]:
        celda.font = negrita
    for item in datos["por_curso"]:
        ws_curso.append([item["curso"], item["cantidad"]])
    ws_curso.column_dimensions["A"].width = 40

    ws_tendencia = wb.create_sheet("Tendencia mensual")
    ws_tendencia.append(["Mes", "Certificados emitidos"])
    for celda in ws_tendencia[1]:
        celda.font = negrita
    for item in datos["tendencia"]:
        ws_tendencia.append([item["mes"].strftime("%m/%Y"), item["cantidad"]])
    ws_tendencia.column_dimensions["A"].width = 15

    if datos["top_usuarios"]:
        ws_usuarios = wb.create_sheet("Certificados por usuario")
        ws_usuarios.append(["Usuario", "Certificados emitidos"])
        for celda in ws_usuarios[1]:
            celda.font = negrita
        for item in datos["top_usuarios"]:
            ws_usuarios.append([item["propietario__username"], item["cantidad"]])
        ws_usuarios.column_dimensions["A"].width = 25

    ws_participantes = wb.create_sheet("Participantes por curso")
    ws_participantes.append(["Curso", "Participantes inscritos"])
    for celda in ws_participantes[1]:
        celda.font = negrita
    for item in datos["por_curso_participantes"]:
        ws_participantes.append([item["lote__curso"], item["cantidad"]])
    ws_participantes.column_dimensions["A"].width = 40

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    respuesta = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    respuesta["Content-Disposition"] = 'attachment; filename="estadisticas_certificados.xlsx"'
    return respuesta


@login_required
def exportar_estadisticas_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    datos = _calcular_estadisticas(request.user)
    estilos = getSampleStyleSheet()
    verde = colors.HexColor("#0d5c3a")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm, leftMargin=1.5 * cm, rightMargin=1.5 * cm,
    )
    elementos = [
        Paragraph("Estadisticas de certificados", estilos["Title"]),
        Paragraph(f'Generado el {timezone.now().strftime("%d/%m/%Y %H:%M")}', estilos["Normal"]),
        Spacer(1, 0.5 * cm),
    ]

    def tabla(encabezados, filas, anchos=None):
        data = [encabezados] + filas
        t = Table(data, colWidths=anchos, hAlign="LEFT")
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), verde),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        return t

    elementos.append(Paragraph("Resumen", estilos["Heading2"]))
    elementos.append(tabla(
        ["Indicador", "Valor"],
        [
            ["Certificados emitidos", str(datos["total_certificados"])],
            ["Participantes inscritos", str(datos["total_participantes"])],
            ["Tipos de certificado (cursos distintos)", str(datos["total_tipos_curso"])],
            ["Lotes generados", str(datos["total_lotes_generados"])],
            ["Total de lotes", str(datos["total_lotes"])],
            ["Promedio de certificados por lote", str(datos["promedio_por_lote"])],
        ],
        anchos=[10 * cm, 6 * cm],
    ))
    elementos.append(Spacer(1, 0.6 * cm))

    elementos.append(Paragraph("Lotes por estado", estilos["Heading2"]))
    elementos.append(tabla(
        ["Estado", "Cantidad", "Porcentaje"],
        [[item["etiqueta"], str(item["cantidad"]), f'{item["porcentaje"]}%'] for item in datos["estados_resumen"]],
        anchos=[6 * cm, 5 * cm, 5 * cm],
    ))
    elementos.append(Spacer(1, 0.6 * cm))

    elementos.append(Paragraph("Certificados emitidos por curso", estilos["Heading2"]))
    if datos["por_curso"]:
        elementos.append(tabla(
            ["Curso", "Certificados emitidos"],
            [[item["curso"], str(item["cantidad"])] for item in datos["por_curso"]],
            anchos=[11 * cm, 5 * cm],
        ))
    else:
        elementos.append(Paragraph("Todavia no hay certificados generados.", estilos["Normal"]))
    elementos.append(Spacer(1, 0.6 * cm))

    elementos.append(Paragraph("Tendencia mensual", estilos["Heading2"]))
    if datos["tendencia"]:
        elementos.append(tabla(
            ["Mes", "Certificados emitidos"],
            [[item["mes"].strftime("%m/%Y"), str(item["cantidad"])] for item in datos["tendencia"]],
            anchos=[6 * cm, 5 * cm],
        ))
    else:
        elementos.append(Paragraph("Todavia no hay certificados generados.", estilos["Normal"]))

    if datos["top_usuarios"]:
        elementos.append(Spacer(1, 0.6 * cm))
        elementos.append(Paragraph("Certificados emitidos por usuario", estilos["Heading2"]))
        elementos.append(tabla(
            ["Usuario", "Certificados emitidos"],
            [[item["propietario__username"], str(item["cantidad"])] for item in datos["top_usuarios"]],
            anchos=[8 * cm, 5 * cm],
        ))

    elementos.append(Spacer(1, 0.6 * cm))
    elementos.append(Paragraph("Participantes inscritos por curso", estilos["Heading2"]))
    if datos["por_curso_participantes"]:
        elementos.append(tabla(
            ["Curso", "Participantes inscritos"],
            [[item["lote__curso"], str(item["cantidad"])] for item in datos["por_curso_participantes"]],
            anchos=[11 * cm, 5 * cm],
        ))
    else:
        elementos.append(Paragraph("Todavia no hay participantes registrados.", estilos["Normal"]))

    doc.build(elementos)
    buffer.seek(0)
    respuesta = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    respuesta["Content-Disposition"] = 'attachment; filename="estadisticas_certificados.pdf"'
    return respuesta


@login_required
def crear_lote(request):
    if request.method == "POST":
        form = LoteForm(request.POST, request.FILES)
        if form.is_valid():
            lote = form.save(commit=False)
            lote.propietario = request.user
            lote.save()
            try:
                poblar_estudiantes_desde_excel(lote)
            except Exception as exc:  # noqa: BLE001 - se informa pero el lote ya quedo creado
                messages.warning(request, f"El lote se creo pero no se pudo leer el Excel: {exc}")
            return redirect("detalle_lote", pk=lote.pk)
    else:
        form = LoteForm()
    return render(request, "certificados/crear_lote.html", {"form": form, "lote": None})


@login_required
def editar_lote(request, pk):
    lote = get_object_or_404(Lote, pk=pk)
    if not _puede_ver(request.user, lote):
        raise Http404

    if request.method == "POST":
        form = LoteForm(request.POST, request.FILES, instance=lote)
        if form.is_valid():
            excel_reemplazado = "excel" in form.changed_data
            form.save()
            if excel_reemplazado:
                try:
                    poblar_estudiantes_desde_excel(lote)
                except Exception as exc:  # noqa: BLE001 - se informa pero el lote ya quedo actualizado
                    messages.warning(request, f"El lote se actualizo pero no se pudo leer el nuevo Excel: {exc}")
            messages.success(request, "Lote actualizado correctamente.")
            return redirect("detalle_lote", pk=lote.pk)
    else:
        form = LoteForm(instance=lote)
    return render(request, "certificados/crear_lote.html", {"form": form, "lote": lote})


@login_required
def detalle_lote(request, pk):
    lote = get_object_or_404(Lote, pk=pk)
    if not _puede_ver(request.user, lote):
        raise Http404
    return render(request, "certificados/detalle_lote.html", {"lote": lote})


@login_required
def eliminar_lote(request, pk):
    lote = get_object_or_404(Lote, pk=pk)
    if not request.user.es_admin:
        raise Http404

    if request.method == "POST":
        curso = lote.curso
        carpeta_lote = Path(settings.MEDIA_ROOT) / "lotes" / str(lote.pk)
        lote.delete()
        if carpeta_lote.exists():
            shutil.rmtree(carpeta_lote, ignore_errors=True)
        messages.success(request, f'Lote "{curso}" eliminado.')
        return redirect("dashboard")

    return render(request, "certificados/eliminar_lote.html", {"lote": lote})


@login_required
def duplicar_lote(request, pk):
    lote = get_object_or_404(Lote, pk=pk)
    if not _puede_ver(request.user, lote):
        raise Http404

    nuevo = Lote(
        propietario=request.user,
        curso=f"{lote.curso} (copia)",
        texto_x=lote.texto_x,
        texto_y=lote.texto_y,
        margen_izquierdo=lote.margen_izquierdo,
        margen_derecho=lote.margen_derecho,
        tamano_fuente=lote.tamano_fuente,
        tamano_fuente_min=lote.tamano_fuente_min,
        color_texto=lote.color_texto,
        fuente=lote.fuente,
    )
    with lote.excel.open("rb") as f:
        nuevo.excel.save(Path(lote.excel.name).name, File(f), save=False)
    with lote.plantilla.open("rb") as f:
        nuevo.plantilla.save(Path(lote.plantilla.name).name, File(f), save=False)
    nuevo.save()

    Estudiante.objects.bulk_create([
        Estudiante(
            lote=nuevo,
            orden=estudiante.orden,
            valor_original=estudiante.valor_original,
            nombre=estudiante.nombre,
            nombre_normalizado=estudiante.nombre_normalizado,
            fecha_registro=estudiante.fecha_registro,
        )
        for estudiante in lote.estudiantes.all()
    ])

    messages.success(request, f'Lote duplicado como "{nuevo.curso}".')
    return redirect("detalle_lote", pk=nuevo.pk)


@login_required
def previsualizar_lote(request, pk):
    lote = get_object_or_404(Lote, pk=pk)
    if not _puede_ver(request.user, lote):
        raise Http404
    try:
        png_bytes = generar_previsualizacion_png(lote)
    except GeneradorError:
        raise Http404
    return HttpResponse(png_bytes, content_type="image/png")


@login_required
def lista_estudiantes(request, pk):
    lote = get_object_or_404(Lote, pk=pk)
    if not _puede_ver(request.user, lote):
        raise Http404
    analisis = analizar_estudiantes(lote, fecha_por_defecto=lote.creado_en.date())

    nombres_con_certificado = set(_mapa_certificados(lote).keys()) if analisis["filas"] else set()

    return render(request, "certificados/lista_estudiantes.html", {
        "lote": lote,
        "analisis": analisis,
        "nombres_con_certificado": nombres_con_certificado,
    })


@login_required
def agregar_estudiantes(request, pk):
    lote = get_object_or_404(Lote, pk=pk)
    if not _puede_ver(request.user, lote):
        raise Http404

    if request.method == "POST":
        texto = request.POST.get("nombres", "")
        nombres_nuevos = [limpiar_nombre(linea) for linea in texto.splitlines()]
        nombres_nuevos = [nombre for nombre in nombres_nuevos if nombre]

        if not nombres_nuevos:
            messages.error(request, "Escribe al menos un nombre para agregar.")
            return redirect("lista_estudiantes", pk=lote.pk)

        hoy = timezone.now().date()
        orden_inicial = (lote.estudiantes.aggregate(maximo=Max("orden"))["maximo"] or 0) + 1
        Estudiante.objects.bulk_create([
            Estudiante(
                lote=lote,
                orden=orden_inicial + indice,
                valor_original=nombre,
                nombre=nombre,
                nombre_normalizado=normalizar_busqueda(nombre),
                fecha_registro=hoy,
            )
            for indice, nombre in enumerate(nombres_nuevos)
        ])
        cantidad = len(nombres_nuevos)
        messages.success(
            request,
            f"Se agregaron {cantidad} participante{'s' if cantidad != 1 else ''} a la lista."
            " Genera de nuevo los certificados para incluirlos.",
        )
        url = reverse("lista_estudiantes", args=[lote.pk])
        return redirect(f"{url}?agregados={cantidad}")

    return redirect("lista_estudiantes", pk=lote.pk)


@login_required
def cambiar_capitalizacion(request, pk):
    lote = get_object_or_404(Lote, pk=pk)
    if not _puede_ver(request.user, lote):
        raise Http404

    if request.method == "POST":
        modo = request.POST.get("modo", "")
        transformaciones = {
            "mayusculas": str.upper,
            "minusculas": str.lower,
            "titulo": str.title,
        }
        transformar = transformaciones.get(modo)
        if not transformar:
            messages.error(request, "Selecciona un formato de nombre valido.")
            return redirect("lista_estudiantes", pk=lote.pk)

        cambiados = 0
        for estudiante in lote.estudiantes.all():
            nombre = limpiar_nombre(estudiante.valor_original)
            if nombre is None:
                continue
            nuevo_valor = transformar(nombre)
            if nuevo_valor != estudiante.valor_original:
                estudiante.valor_original = nuevo_valor
                estudiante.nombre = nuevo_valor
                estudiante.save(update_fields=["valor_original", "nombre", "nombre_normalizado"])
                cambiados += 1

        messages.success(
            request,
            f"Se actualizo el formato de {cambiados} nombre{'s' if cambiados != 1 else ''}."
            " Genera de nuevo los certificados para reflejar el cambio.",
        )

    return redirect("lista_estudiantes", pk=lote.pk)


@login_required
def editar_estudiante(request, pk, fila):
    lote = get_object_or_404(Lote, pk=pk)
    if not _puede_ver(request.user, lote):
        raise Http404

    if request.method == "POST":
        nuevo_valor = limpiar_nombre(request.POST.get("valor", ""))
        if not nuevo_valor:
            messages.error(request, "El nombre no puede quedar vacio. Usa 'Eliminar' si deseas quitarlo.")
        else:
            estudiante = get_object_or_404(Estudiante, lote=lote, orden=fila)
            estudiante.valor_original = nuevo_valor
            estudiante.nombre = nuevo_valor
            estudiante.save(update_fields=["valor_original", "nombre", "nombre_normalizado"])
            messages.success(request, "Participante actualizado. Genera de nuevo los certificados para reflejar el cambio.")

    return redirect("lista_estudiantes", pk=lote.pk)


@login_required
def eliminar_estudiante(request, pk, fila):
    lote = get_object_or_404(Lote, pk=pk)
    if not _puede_ver(request.user, lote):
        raise Http404

    if request.method == "POST":
        estudiante = get_object_or_404(Estudiante, lote=lote, orden=fila)
        estudiante.delete()
        messages.success(request, "Participante eliminado. Genera de nuevo los certificados para reflejar el cambio.")

    return redirect("lista_estudiantes", pk=lote.pk)


@login_required
def descargar_certificado_individual(request, pk, fila):
    lote = get_object_or_404(Lote, pk=pk)
    if not _puede_ver(request.user, lote):
        raise Http404

    analisis = analizar_estudiantes(lote, fecha_por_defecto=lote.creado_en.date())

    item = next((f for f in analisis["filas"] if f["fila"] == fila), None)
    if not item or not item["nombre"]:
        raise Http404

    ruta = _mapa_certificados(lote).get(item["nombre"])
    if not ruta:
        messages.error(request, "Ese certificado aun no esta generado. Genera (o regenera) el lote primero.")
        return redirect("lista_estudiantes", pk=lote.pk)

    return FileResponse(ruta.open("rb"), as_attachment=True, filename=ruta.name)


@login_required
def generar_certificado_individual(request, pk, fila):
    lote = get_object_or_404(Lote, pk=pk)
    if not _puede_ver(request.user, lote):
        raise Http404
    if request.method != "POST":
        raise Http404

    analisis = analizar_estudiantes(lote, fecha_por_defecto=lote.creado_en.date())
    item = next((f for f in analisis["filas"] if f["fila"] == fila), None)
    if not item or not item["nombre"]:
        raise Http404

    carpeta_salida = Path(settings.MEDIA_ROOT) / "lotes" / str(lote.pk) / "salida"
    try:
        generar_certificado_individual_pdf(lote, item["nombre"], carpeta_salida)
    except GeneradorError as exc:
        messages.error(request, f"No se pudo generar el certificado: {exc}")
    else:
        messages.success(request, f'Certificado de "{item["nombre"]}" generado correctamente.')

    return redirect("lista_estudiantes", pk=lote.pk)


@login_required
def descargar_certificados(request, pk):
    lote = get_object_or_404(Lote, pk=pk)
    if not _puede_ver(request.user, lote):
        raise Http404
    if request.method != "POST":
        raise Http404

    filas_seleccionadas = request.POST.getlist("fila")
    if not filas_seleccionadas:
        messages.error(request, "Selecciona al menos un participante para descargar.")
        return redirect("lista_estudiantes", pk=lote.pk)

    analisis = analizar_estudiantes(lote, fecha_por_defecto=lote.creado_en.date())

    filas_por_numero = {str(item["fila"]): item for item in analisis["filas"]}
    mapa_certificados = _mapa_certificados(lote)

    rutas = []
    faltantes = []
    vistos = set()
    for fila in filas_seleccionadas:
        item = filas_por_numero.get(fila)
        nombre = item["nombre"] if item else None
        if not nombre or nombre in vistos:
            continue
        vistos.add(nombre)
        ruta = mapa_certificados.get(nombre)
        if ruta:
            rutas.append((nombre, ruta))
        else:
            faltantes.append(nombre)

    if not rutas:
        messages.error(
            request,
            "No hay certificados generados para los participantes seleccionados. Genera (o regenera) el lote primero.",
        )
        return redirect("lista_estudiantes", pk=lote.pk)

    if faltantes:
        messages.warning(request, f"No se encontro certificado generado para: {', '.join(faltantes)}.")

    if len(rutas) == 1:
        nombre, ruta = rutas[0]
        return FileResponse(ruta.open("rb"), as_attachment=True, filename=ruta.name)

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for _, ruta in rutas:
            zf.write(ruta, arcname=ruta.name)
    buffer.seek(0)
    respuesta = HttpResponse(buffer.getvalue(), content_type="application/zip")
    respuesta["Content-Disposition"] = f'attachment; filename="certificados_{lote.curso}_seleccion.zip"'
    return respuesta


@login_required
def exportar_reporte(request, pk):
    lote = get_object_or_404(Lote, pk=pk)
    if not _puede_ver(request.user, lote):
        raise Http404
    analisis = analizar_estudiantes(lote, fecha_por_defecto=lote.creado_en.date())
    contenido = exportar_reporte_excel(analisis)
    respuesta = HttpResponse(contenido, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    respuesta["Content-Disposition"] = f'attachment; filename="reporte_{lote.curso}.xlsx"'
    return respuesta


@login_required
def exportar_paquete_curso(request, pk):
    lote = get_object_or_404(Lote, pk=pk)
    if not _puede_ver(request.user, lote):
        raise Http404

    analisis = analizar_estudiantes(lote, fecha_por_defecto=lote.creado_en.date())
    reporte_excel = exportar_reporte_excel(analisis)

    nombre_base = nombre_archivo_seguro(lote.curso)

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"curso.txt", lote.curso)
        zf.writestr(f"participantes_{nombre_base}.xlsx", reporte_excel)
        if lote.plantilla:
            try:
                with lote.plantilla.open("rb") as f:
                    extension = Path(lote.plantilla.name).suffix or ".png"
                    zf.writestr(f"plantilla_{nombre_base}{extension}", f.read())
            except FileNotFoundError:
                messages.warning(request, "La imagen de plantilla no se encontro en el servidor; el paquete se genero sin ella.")
    buffer.seek(0)

    respuesta = HttpResponse(buffer.getvalue(), content_type="application/zip")
    respuesta["Content-Disposition"] = f'attachment; filename="curso_{nombre_base}.zip"'
    return respuesta


def _generar_en_segundo_plano(lote_id):
    close_old_connections()
    lote = Lote.objects.get(pk=lote_id)

    def reportar(procesados, total):
        Lote.objects.filter(pk=lote_id).update(procesados=procesados, total_certificados=total)

    try:
        carpeta_salida = Path(settings.MEDIA_ROOT) / "lotes" / str(lote.pk) / "salida"
        if carpeta_salida.exists():
            shutil.rmtree(carpeta_salida, ignore_errors=True)
        cantidad, zip_path = generar_certificados_lote(lote, carpeta_salida, on_progreso=reportar)
        lote.refresh_from_db()
        lote.zip_resultado = zip_path.relative_to(settings.MEDIA_ROOT).as_posix()
        lote.total_certificados = cantidad
        lote.procesados = cantidad
        lote.archivos_expirados = False
        lote.estado = Lote.Estado.GENERADO
        lote.mensaje_error = ""
        lote.save()
    except GeneradorError as exc:
        Lote.objects.filter(pk=lote_id).update(estado=Lote.Estado.ERROR, mensaje_error=str(exc))
    except Exception as exc:  # noqa: BLE001 - se reporta como error de lote, no debe tumbar el hilo
        Lote.objects.filter(pk=lote_id).update(estado=Lote.Estado.ERROR, mensaje_error=f"Error inesperado: {exc}")
    finally:
        close_old_connections()


@login_required
def generar_lote(request, pk):
    lote = get_object_or_404(Lote, pk=pk)
    if not _puede_ver(request.user, lote):
        raise Http404

    lote.estado = Lote.Estado.PROCESANDO
    lote.procesados = 0
    lote.total_certificados = 0
    lote.mensaje_error = ""
    lote.generado_por = request.user
    lote.generado_en = timezone.now()
    lote.save()

    hilo = threading.Thread(target=_generar_en_segundo_plano, args=(lote.pk,), daemon=True)
    hilo.start()

    return redirect("detalle_lote", pk=lote.pk)


@login_required
def progreso_lote(request, pk):
    lote = get_object_or_404(Lote, pk=pk)
    if not _puede_ver(request.user, lote):
        raise Http404
    return JsonResponse({
        "estado": lote.estado,
        "procesados": lote.procesados,
        "total": lote.total_certificados,
        "mensaje_error": lote.mensaje_error,
        "tiene_zip": bool(lote.zip_resultado),
    })


@login_required
def descargar_zip(request, pk):
    lote = get_object_or_404(Lote, pk=pk)
    if not _puede_ver(request.user, lote):
        raise Http404
    if not lote.zip_resultado:
        raise Http404
    return FileResponse(
        lote.zip_resultado.open("rb"),
        as_attachment=True,
        filename=f"certificados_{lote.curso}.zip",
    )
